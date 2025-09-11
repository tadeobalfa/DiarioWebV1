# app.py
# ------------------------------------------------------------
# Versi√≥n Estable Diario V1 + Cierres + Mayor + Filtro de Ajustes por color
# - Mantiene TODO lo aprobado (no se cambia nada salvo lo pedido)
# - Ajustes: solo se incluyen bloques cuyo n√∫mero est√° pintado AMARILLO/NARANJA.
#   Si est√° sin color o ROJO, no se incluye ni se arrastran l√≠neas.
# - Cierres (Resultado y Patrimonial) con ‚ÄúResultado del Ejercicio‚Äù
# - Mayor al final (Cuenta / Total Debe / Total Haber) tras 3 filas en blanco
# - En UI: selecci√≥n expl√≠cita de la columna de saldos reexpresados (por letra o nombre)
# ------------------------------------------------------------

import io
import re
import math
import unicodedata
import calendar
from datetime import datetime, date
from typing import List, Tuple, Optional, Dict

import pandas as pd
import streamlit as st

# ==========================
# Config Streamlit
# ==========================
st.set_page_config(page_title="Generador de Diario", layout="wide")
st.title("üìò Generador de Asientos Diarios")
st.caption("Versi√≥n Estable Diario V1 ‚Äî Apertura + 12 meses + Ajustes + Cierres + Mayor")

# ==========================
# Utilidades generales
# ==========================

def _strip_accents_lower(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return " ".join(s.lower().split())

BANNED_SECTION_LABELS = {
    "activo", "pasivo", "patrimonio neto", "ingresos", "egresos", "total", "totales"
}

def is_banned_label(texto: str) -> bool:
    return _strip_accents_lower(texto) in BANNED_SECTION_LABELS

OPENING_COL_SYNS = {
    "saldo al inicio", "saldo inicial", "saldo inicio", "saldo de apertura", "saldo apertura"
}

MONTHS_ES = {
    1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
    7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
}

MONTH_HEADER_PAT = re.compile(
    r"(?i)(?:\b(enero|ene)\b|\b(febrero|feb)\b|\b(marzo|mar)\b|\b(abril|abr)\b|\b(mayo|may)\b|\b(junio|jun)\b|\b(julio|jul)\b|\b(agosto|ago)\b|\b(septiembre|setiembre|sep|set)\b|\b(octubre|oct)\b|\b(noviembre|nov)\b|\b(diciembre|dic)\b)"
)

def _month_number_from_header(h: str) -> Optional[int]:
    if h is None:
        return None
    s = _strip_accents_lower(h)

    m = MONTH_HEADER_PAT.search(s)
    if m:
        group_to_month = {1:1, 2:2, 3:3, 4:4, 5:5, 6:6, 7:7, 8:8, 9:9, 10:10, 11:11, 12:12}
        for idx, month_num in group_to_month.items():
            if m.group(idx):
                return month_num

    m = re.search(r"\b(20\d{2})[-_/](0?[1-9]|1[0-2])\b", s)
    if m:
        return int(m.group(2))
    m = re.search(r"\b(0?[1-9]|1[0-2])[-_/ ](20\d{2}|\d{2})\b", s)
    if m:
        return int(m.group(1))
    if re.fullmatch(r"(0?[1-9]|1[0-2])", s.replace(" ", "")):
        if not any(t in s for t in ("debe","haber","saldo","apertura","inicial","inicio","total")):
            return int(s)
    return None

def _year_from_header(h: str) -> Optional[int]:
    if h is None:
        return None
    s = str(h)
    m = re.search(r"\b(20\d{2})\b", s)
    if m:
        return int(m.group(1))
    try:
        dt = pd.to_datetime(s, errors="coerce", dayfirst=False)
        if pd.notna(dt):
            return int(dt.year)
    except Exception:
        pass
    return None

def _looks_like_month(colname: str) -> bool:
    return _month_number_from_header(str(colname)) is not None

def find_header_row(df_raw: pd.DataFrame, max_scan_rows: int = 60) -> int:
    best_row = 0
    best_hits = -1
    for r in range(min(max_scan_rows, len(df_raw))):
        hits = sum(_looks_like_month(v) for v in df_raw.iloc[r].tolist())
        if hits > best_hits:
            best_hits = hits
            best_row = r
    return best_row

def _first_non_empty_in_first_col(df_raw: pd.DataFrame) -> Optional[str]:
    try:
        col0 = df_raw.iloc[:,0]
        for v in col0:
            if pd.notna(v) and str(v).strip():
                return str(v).strip()
    except Exception:
        pass
    return None

def _infer_year_from_headers(headers: List[str]) -> Optional[int]:
    years = []
    for h in headers:
        y = _year_from_header(h)
        if y:
            years.append(y)
    if years:
        return max(set(years), key=years.count)
    return None

# ==========================
# Carga del balance
# ==========================

def load_balance_from_bytes(data: bytes, sheet_name: str):
    """
    Devuelve:
      df, header_row, opening_col, ordered_months [(y,m,col)], account_col,
      empresa_raw, period_start (y,m), period_end (y,m)
    """
    buff = io.BytesIO(data)
    df_raw = pd.read_excel(buff, sheet_name=sheet_name, header=None, dtype=object)
    header_row = find_header_row(df_raw)

    buff.seek(0)
    df = pd.read_excel(buff, sheet_name=sheet_name, header=header_row, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    all_cols = list(df.columns)
    months_like = [c for c in all_cols if _looks_like_month(c)]
    non_month = [c for c in all_cols if c not in months_like]
    account_col = non_month[0] if non_month else all_cols[0]

    opening_col = None
    for c in all_cols:
        if _strip_accents_lower(c) in OPENING_COL_SYNS:
            opening_col = c; break
    if opening_col is None:
        for c in all_cols:
            s = _strip_accents_lower(c)
            if "saldo" in s and any(t in s for t in ("inicio","inicial","apertura")):
                opening_col = c; break

    month_tuples: List[Tuple[int,int,str]] = []
    fallback_year = _infer_year_from_headers(all_cols)
    for c in all_cols:
        mnum = _month_number_from_header(c)
        if not mnum:
            continue
        y = _year_from_header(c) or fallback_year or datetime.now().year
        month_tuples.append((y, mnum, c))

    month_tuples.sort(key=lambda t: (t[0], t[1]))
    seen = set()
    ordered_months: List[Tuple[int,int,str]] = []
    for y, m, col in month_tuples:
        key = (y, m)
        if key not in seen:
            seen.add(key)
            ordered_months.append((y, m, col))
    if len(ordered_months) > 12:
        ordered_months = ordered_months[:12]

    if ordered_months:
        period_start = (ordered_months[0][0], ordered_months[0][1])
        period_end   = (ordered_months[-1][0], ordered_months[-1][1])
    else:
        y = fallback_year or datetime.now().year
        period_start = (y, 1)
        period_end   = (y, 12)

    empresa_raw = _first_non_empty_in_first_col(df_raw)

    return df, header_row, opening_col, ordered_months, account_col, empresa_raw, period_start, period_end

# ==========================
# Parsing num√©rico robusto
# ==========================

def _to_float(x) -> float:
    if x is None: return 0.0
    if isinstance(x, (int, float)):
        try:
            fx = float(x)
            return fx if math.isfinite(fx) else 0.0
        except Exception:
            return 0.0

    s = str(x).strip()
    if not s: return 0.0
    s = s.replace("\u00A0", " ").replace(" ", "")
    for sym in "$‚Ç¨¬£ARSAR$":
        s = s.replace(sym, "")

    has_dot = "." in s
    has_comma = "," in s
    try:
        if has_dot and has_comma:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
                return float(s)
            else:
                s = s.replace(",", "")
                return float(s)
        elif has_comma and not has_dot:
            if s.count(",") == 1:
                s = s.replace(",", "."); return float(s)
            else:
                s = s.replace(",", "");  return float(s)
        elif has_dot and not has_comma:
            if s.count(".") > 1:
                s = s.replace(".", "");  return float(s)
            else:
                return float(s)
        else:
            return float(s)
    except Exception:
        return 0.0

def _safe_num(x) -> float:
    v = _to_float(x)
    return v if math.isfinite(v) else 0.0

# ==========================
# Construcci√≥n de l√≠neas (base estable)
# ==========================

def build_opening_lines(df: pd.DataFrame, opening_col: Optional[str], account_col: str) -> List[dict]:
    if not opening_col or opening_col not in df.columns: return []
    lines = []
    for _, row in df.iterrows():
        cuenta = str(row.get(account_col, "")).strip()
        if not cuenta or is_banned_label(cuenta): continue
        val = _safe_num(row.get(opening_col, 0))
        if abs(val) < 1e-9: continue
        if val >= 0:
            lines.append({"Cuenta": cuenta, "Debe": abs(val), "Haber": 0.0})
        else:
            lines.append({"Cuenta": cuenta, "Debe": 0.0, "Haber": abs(val)})
    return lines

def build_month_lines(df: pd.DataFrame, month_col: Optional[str], account_col: str) -> List[dict]:
    if not month_col or month_col not in df.columns: return []
    lines = []
    for _, row in df.iterrows():
        cuenta = str(row.get(account_col, "")).strip()
        if not cuenta or is_banned_label(cuenta): continue
        val = _safe_num(row.get(month_col, 0))
        if abs(val) < 1e-9: continue
        if val >= 0:
            lines.append({"Cuenta": cuenta, "Debe": abs(val), "Haber": 0.0})
        else:
            lines.append({"Cuenta": cuenta, "Debe": 0.0, "Haber": abs(val)})
    return lines

# ==========================
# AJUSTES (solo bloques AMARILLO/NARANJA)
# ==========================

YELLOW_ORANGE_HEXES = {
    "FFFFFF00","FFFFF200","FFFFE699","FFFFF2CC","FFFFEB84","FFFFF4B6","FFFFF9CC",
    "FFFFC000","FFF4B183","FFF8CBAD","FFFFD965","FFFFB84D"
}
YELLOW_ORANGE_RGB = {h[-6:] for h in YELLOW_ORANGE_HEXES}

def _normalize_hex_from_cell(cell) -> Optional[str]:
    def extract_hex(color_obj):
        if color_obj is None:
            return None
        for attr in ("rgb", "value"):
            v = getattr(color_obj, attr, None)
            if v is None: 
                continue
            v = getattr(v, "rgb", v)
            v = getattr(v, "value", v)
            s = str(v).upper()
            hexchars = "".join(ch for ch in s if ch in "0123456789ABCDEF")
            if len(hexchars) >= 6:
                return ("FF" + hexchars[-6:]) if len(hexchars) == 6 else hexchars
        return None

    for color in (
        getattr(cell.fill, "fgColor", None),
        getattr(cell.fill, "start_color", None),
        getattr(cell.fill, "end_color", None),
    ):
        hx = extract_hex(color)
        if hx:
            return hx
    return None

def _get_top_left_cell(ws, cell):
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return cell

def _is_numeric_like(val) -> bool:
    try:
        if val is None: return False
        float(str(val).strip().replace(".", "").replace(",", "."))
        return True
    except Exception:
        return False

def _marker_kind(ws, row_idx: int) -> str:
    """
    Devuelve:
      - 'include' si el n√∫mero de asiento (A o B) est√° pintado amarillo/naranja
      - 'exclude' si hay n√∫mero de asiento sin color o con otro color (rojo, etc.)
      - 'none'    si no hay n√∫mero de asiento en la fila
    """
    for col in (1, 2):  # A o B
        cell = ws.cell(row_idx, col)
        tl = _get_top_left_cell(ws, cell)
        val = cell.value if cell.value is not None else tl.value

        if not _is_numeric_like(val):
            continue

        hx1 = _normalize_hex_from_cell(tl)
        hx2 = _normalize_hex_from_cell(cell)
        hx = (hx1 or hx2) or ""   # puede ser None si no tiene color

        if (hx in YELLOW_ORANGE_HEXES) or (hx[-6:] in YELLOW_ORANGE_RGB):
            return 'include'
        else:
            return 'exclude'
    return 'none'

def read_adjust_blocks_from_bytes(data: bytes, sheet_name: Optional[str]) -> List[Tuple[str, List[dict]]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)

    ws = None
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for c in ["As Aj", "AS AJ", "AS AJ.", "AS AJUSTE", "AS AJUSTES"]:
            if c in wb.sheetnames:
                ws = wb[c]; break
        if ws is None:
            return []

    blocks: List[Tuple[str, List[dict]]] = []
    current: List[dict] = []
    title: Optional[str] = None
    include_mode = False  # solo capturamos si el marcador fue 'include'

    for r in range(2, ws.max_row + 1):
        kind = _marker_kind(ws, r)

        if kind in ('include', 'exclude'):
            if include_mode and current:
                blocks.append((title or "Asientos de Ajuste", current))
            current = []
            title = None
            include_mode = (kind == 'include')
            continue

        if not include_mode:
            continue

        cuenta = ws.cell(r, 1).value
        nombre = ws.cell(r, 2).value
        debe   = ws.cell(r, 3).value
        haber  = ws.cell(r, 4).value

        if (cuenta is None and nombre is None and debe is None and haber is None):
            continue

        if title is None and nombre not in (None, ""):
            title = str(nombre).strip()

        if (cuenta is None or str(cuenta).strip() == "") and (debe in (None, "")) and (haber in (None, "")):
            continue

        current.append({
            "Cuenta": str(cuenta or "").strip(),
            "Debe": _safe_num(debe),
            "Haber": _safe_num(haber)
        })

    if include_mode and current:
        blocks.append((title or "Asientos de Ajuste", current))

    return blocks

# ==========================
# Fechas
# ==========================

def _fmt_dmy(d: date) -> str:
    return d.strftime("%d/%m/%Y")

def first_day(y: int, m: int) -> date:
    return date(y, m, 1)

def last_day(y: int, m: int) -> date:
    return date(y, m, calendar.monthrange(y, m)[1])

# ==========================
# Exportar a Excel (incluye MAYOR)
# ==========================

def write_entry(ws, start_row: int, asiento_num: int, titulo: str, lines: List[dict], workbook, fecha_str: str):
    header_fmt = workbook.add_format({"bold": True, "bottom": 1})
    money_fmt  = workbook.add_format({"num_format": '#,##0.00'})
    total_fmt  = workbook.add_format({"bold": True, "top": 1, "num_format": '#,##0.00'})
    bold_fmt   = workbook.add_format({"bold": True})

    row = start_row
    # Encabezado por asiento: A=Fecha, B=T√≠tulo, C=Nro asiento, D vac√≠o, E=Debe, F=Haber
    ws.write(row, 0, "Fecha", header_fmt)
    ws.write(row, 1, titulo, header_fmt)
    ws.write(row, 2, asiento_num, header_fmt)
    ws.write(row, 3, "", header_fmt)
    ws.write(row, 4, "Debe", header_fmt)
    ws.write(row, 5, "Haber", header_fmt)
    row += 1

    total_debe = 0.0
    total_haber = 0.0
    printed_first = False

    for ln in lines:
        cuenta = ln.get("Cuenta","")
        debe   = _safe_num(ln.get("Debe",0))
        haber  = _safe_num(ln.get("Haber",0))

        ws.write(row, 0, fecha_str if not printed_first else "")
        ws.write(row, 2, cuenta)
        ws.write_number(row, 4, debe, money_fmt)
        ws.write_number(row, 5, haber, money_fmt)

        total_debe += debe; total_haber += haber
        printed_first = True
        row += 1

    if not printed_first:
        ws.write(row, 0, fecha_str)
        row += 1

    # Fila en blanco + fila con el t√≠tulo en C (negrita) antes de totales
    row += 1
    ws.write(row, 2, titulo, bold_fmt)
    row += 1

    # Totales solo n√∫meros
    ws.write_number(row, 4, total_debe, total_fmt)
    ws.write_number(row, 5, total_haber, total_fmt)
    row += 3  # dos filas en blanco

    return row

def write_blue_separator(ws, start_row: int, workbook, cols=6):
    blue = workbook.add_format({"bg_color": "#CFE8FF"})
    r = start_row
    for c in range(cols):
        ws.write(r, c, "", blue)
    return r + 1

def _accumulate_major(agg: Dict[str, Dict[str, float]], lines: List[dict]):
    for ln in lines:
        cta = str(ln.get("Cuenta","")).strip()
        if not cta:
            continue
        d = _safe_num(ln.get("Debe",0.0))
        h = _safe_num(ln.get("Haber",0.0))
        if cta not in agg:
            agg[cta] = {"Debe": 0.0, "Haber": 0.0}
        agg[cta]["Debe"]  += d
        agg[cta]["Haber"] += h

def _write_mayor(ws, start_row: int, workbook, mayor_agg: Dict[str, Dict[str, float]]):
    bold = workbook.add_format({"bold": True})
    money = workbook.add_format({"num_format": '#,##0.00'})

    row = start_row
    ws.write(row, 0, "Cuenta", bold)
    ws.write(row, 1, "Debe", bold)
    ws.write(row, 2, "Haber", bold)
    row += 1

    for cta in sorted(mayor_agg.keys()):
        ws.write(row, 0, cta)
        ws.write_number(row, 1, _safe_num(mayor_agg[cta]["Debe"]), money)
        ws.write_number(row, 2, _safe_num(mayor_agg[cta]["Haber"]), money)
        row += 1

    return row

def build_output_excel(empresa: str,
                       opening_tuple: Tuple[str, List[dict], str],
                       month_tuples: List[Tuple[str, List[dict], str]],
                       adjust_blocks: List[Tuple[str, List[dict]]],
                       cierre_resultado: Optional[List[dict]],
                       cierre_patrimonial: Optional[List[dict]],
                       period_end_date: date) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(
        output,
        engine="xlsxwriter",
        engine_kwargs={"options": {"nan_inf_to_errors": True}}
    ) as writer:
        wb = writer.book
        ws = wb.add_worksheet("DIARIO")

        ws.set_column(0, 0, 16)  # A (Fecha en asientos / Cuenta en Mayor)
        ws.set_column(1, 1, 28)  # B T√≠tulo
        ws.set_column(2, 2, 36)  # C Cuenta
        ws.set_column(3, 3, 6)   # D vac√≠o
        ws.set_column(4, 5, 14)  # E-F Debe/Haber

        row = 0
        ws.write(row, 0, empresa or "EMPRESA"); row += 1
        ws.write(row, 0, f"ASIENTOS DIARIOS - {period_end_date.year}"); row += 1
        row += 1

        asiento = 1
        mayor_agg: Dict[str, Dict[str, float]] = {}

        # Apertura
        t_open, lines_open, f_open = opening_tuple
        row = write_entry(ws, row, asiento, t_open, lines_open, wb, f_open); asiento += 1
        _accumulate_major(mayor_agg, lines_open)

        # Mensuales
        for t, lines, f in month_tuples:
            row = write_entry(ws, row, asiento, t, lines, wb, f); asiento += 1
            _accumulate_major(mayor_agg, lines)

        # Ajustes
        if any(lines for _, lines in adjust_blocks):
            row = write_blue_separator(ws, row, wb)
            fecha_aj = _fmt_dmy(period_end_date)
            for title, lines in adjust_blocks:
                if not lines: continue
                row = write_entry(ws, row, asiento, title or "Asientos de Ajuste", lines, wb, fecha_aj)
                asiento += 1
                _accumulate_major(mayor_agg, lines)

        # Cierres (misma fecha que ajustes)
        fecha_cierre = _fmt_dmy(period_end_date)
        if cierre_resultado:
            titulo_res = f"Cierre de Cuentas de Resultado {period_end_date.year}"
            row = write_entry(ws, row, asiento, titulo_res, cierre_resultado, wb, fecha_cierre)
            asiento += 1
            _accumulate_major(mayor_agg, cierre_resultado)
        if cierre_patrimonial:
            titulo_pat = f"Cierre de Cuentas Patrimoniales {period_end_date.year}"
            row = write_entry(ws, row, asiento, titulo_pat, cierre_patrimonial, wb, fecha_cierre)
            asiento += 1
            _accumulate_major(mayor_agg, cierre_patrimonial)

        # Mayor (3 filas en blanco antes)
        row += 3
        _write_mayor(ws, row, wb, mayor_agg)

        writer.close()
    return output.getvalue()

# ==========================
# CIERRE: armado a partir de columna reexpresada
# ==========================

def _col_letter_to_index(letter: str) -> Optional[int]:
    s = str(letter).strip().upper()
    if not s or not s.isalpha():
        return None
    idx = 0
    for ch in s:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1

def _pick_reexpresado_column(df: pd.DataFrame, user_choice: str) -> Optional[str]:
    if not user_choice:
        return None
    if user_choice in df.columns:
        return user_choice
    col_idx = _col_letter_to_index(user_choice)
    if col_idx is not None and 0 <= col_idx < len(df.columns):
        return df.columns[col_idx]
    return None

def _build_cierre_from_df(df: pd.DataFrame, account_col: str, reexp_col: str) -> Tuple[List[dict], List[dict]]:
    df2 = df[[account_col, reexp_col]].copy()
    df2.columns = [account_col, "SALDO_U"]
    df2 = df2.dropna(subset=[account_col], how="all")
    df2[account_col] = df2[account_col].astype(str).str.strip()

    idx_ingresos = df2[df2[account_col].str.upper().str.fullmatch("INGRESOS")].index
    idx_totales  = df2[df2[account_col].str.upper().str.startswith("TOTA")].index
    start_ingresos = idx_ingresos[0] if len(idx_ingresos)>0 else None
    start_totales  = idx_totales[0]  if len(idx_totales)>0 else len(df2)

    df_patr = df2.loc[:start_ingresos-1] if start_ingresos is not None else df2.copy()
    df_resu = df2.loc[start_ingresos+1:start_totales-1] if start_ingresos is not None else pd.DataFrame(columns=df2.columns)

    banned = {s.upper() for s in BANNED_SECTION_LABELS}

    res_lines: List[dict] = []
    for _, row in df_resu.iterrows():
        cta = str(row[account_col]).strip()
        if cta.upper() in banned: continue
        v = _safe_num(row["SALDO_U"])
        if abs(v) < 1e-9: continue
        if v > 0:
            res_lines.append({"Cuenta": cta, "Debe": 0.0, "Haber": v})
        else:
            res_lines.append({"Cuenta": cta, "Debe": -v, "Haber": 0.0})

    pat_lines: List[dict] = []
    for _, row in df_patr.iterrows():
        cta = str(row[account_col]).strip()
        if cta.upper() in banned: continue
        v = _safe_num(row["SALDO_U"])
        if abs(v) < 1e-9: continue
        if v > 0:
            pat_lines.append({"Cuenta": cta, "Debe": 0.0, "Haber": v})
        else:
            pat_lines.append({"Cuenta": cta, "Debe": -v, "Haber": 0.0})

    return res_lines, pat_lines

def _add_resultado_del_ejercicio(res_lines: List[dict], pat_lines: List[dict]) -> Tuple[List[dict], List[dict]]:
    def sum_dh(lines):
        d = sum(_safe_num(x["Debe"]) for x in lines)
        h = sum(_safe_num(x["Haber"]) for x in lines)
        return d, h

    d_res, h_res = sum_dh(res_lines)
    diff = d_res - h_res
    if diff > 0:
        res_lines.append({"Cuenta": "Resultado del Ejercicio", "Debe": 0.0, "Haber": diff})
        pat_lines.append({"Cuenta": "Resultado del Ejercicio", "Debe": diff, "Haber": 0.0})
    elif diff < 0:
        res_lines.append({"Cuenta": "Resultado del Ejercicio", "Debe": -diff, "Haber": 0.0})
        pat_lines.append({"Cuenta": "Resultado del Ejercicio", "Debe": 0.0, "Haber": -diff})
    return res_lines, pat_lines

# ==========================
# UI principal
# ==========================

st.subheader("1) Cargar archivo y elegir hojas")
uploaded = st.file_uploader("Sub√≠ tu Excel (.xlsx)", type=["xlsx"])

if uploaded:
    excel_bytes = uploaded.read()
    xls = pd.ExcelFile(io.BytesIO(excel_bytes))
    sheet_names = xls.sheet_names

    col1, col2 = st.columns(2)
    with col1:
        hoja_balance = st.selectbox("Hoja del BALANCE", sheet_names, index=0)
    with col2:
        hoja_ajustes = st.selectbox("Hoja de AJUSTES (opcional)", ["Ninguna"] + sheet_names, index=0)

    # Selector columna reexpresada
    st.subheader("2) Indicar columna de SALDOS REEXPRESADOS")
    detected_cols = []
    try:
        df_tmp_raw = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=hoja_balance, header=None)
        hdr = find_header_row(df_tmp_raw)
        df_hdr = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=hoja_balance, header=hdr, nrows=0)
        detected_cols = list(df_hdr.columns)
    except Exception:
        pass

    select_options = []
    for i, c in enumerate(detected_cols):
        n = i
        letters = ""
        while True:
            n, r = divmod(n, 26)
            letters = chr(65 + r) + letters
            if n == 0:
                break
        select_options.append(f"{letters} ‚Äî {c}")

    default_idx = 20 if len(select_options) > 20 else 0
    choice = st.selectbox(
        "Seleccion√° la columna con los SALDOS REEXPRESADOS (por letra o nombre de encabezado):",
        options=select_options + ["Otra (ingresar manualmente)"],
        index=default_idx if default_idx < len(select_options) else len(select_options)
    )
    manual_input = ""
    if choice == "Otra (ingresar manualmente)":
        manual_input = st.text_input("Ingres√° la **letra de columna** (ej. U) o el **nombre exacto** del encabezado:")

    try:
        # Balance base
        (df_bal, header_row, opening_col, ordered_months,
         account_col, empresa_raw, period_start, period_end) = load_balance_from_bytes(excel_bytes, hoja_balance)

        empresa = empresa_raw or uploaded.name.rsplit(".",1)[0]

        with st.expander("üîé Diagn√≥stico"):
            st.write(f"**Fila de encabezado detectada:** {header_row}")
            st.write(f"**Columna de APERTURA:** {opening_col}")
            st.write(f"**Columna de CUENTA/CONCEPTO:** {account_col}")
            st.write("**Meses detectados (ordenados):**", [(y, m, c) for (y,m,c) in ordered_months])
            st.write(f"**Periodo:** {period_start[1]:02d}/{period_start[0]} ‚Üí {period_end[1]:02d}/{period_end[0]}")

        # L√≠neas y fechas
        opening_lines = build_opening_lines(df_bal, opening_col, account_col)
        open_date = first_day(period_start[0], period_start[1])
        opening_title = f"Asiento de Apertura {open_date.year}"
        opening_tuple = (opening_title, opening_lines, _fmt_dmy(open_date))

        month_tuples: List[Tuple[str, List[dict], str]] = []
        y, m = period_start
        cols_map = {(yy, mm): col for (yy, mm, col) in ordered_months}
        for _ in range(12):
            yy, mm = y, m
            col = cols_map.get((yy, mm))
            label = f"As. Movimiento {MONTHS_ES[mm]} {yy}"
            lines = build_month_lines(df_bal, col, account_col) if col else []
            fecha = _fmt_dmy(last_day(yy, mm))
            month_tuples.append((label, lines, fecha))
            m += 1
            if m == 13:
                y += 1; m = 1

        # Ajustes
        adjust_blocks: List[Tuple[str, List[dict]]] = []
        if hoja_ajustes != "Ninguna":
            adjust_blocks = read_adjust_blocks_from_bytes(excel_bytes, sheet_name=hoja_ajustes)

        # Cierres
        if choice != "Otra (ingresar manualmente)":
            sel_index = select_options.index(choice)
            real_col = detected_cols[sel_index] if sel_index < len(detected_cols) else None
        else:
            real_col = _pick_reexpresado_column(
                pd.read_excel(io.BytesIO(excel_bytes), sheet_name=hoja_balance, header=header_row),
                manual_input
            )

        cierre_resultado_lines: List[dict] = []
        cierre_patrimonial_lines: List[dict] = []
        if real_col and account_col in df_bal.columns and real_col in df_bal.columns:
            res_lines, pat_lines = _build_cierre_from_df(df_bal, account_col, real_col)
            res_lines, pat_lines = _add_resultado_del_ejercicio(res_lines, pat_lines)
            cierre_resultado_lines = res_lines
            cierre_patrimonial_lines = pat_lines

        period_end_date = last_day(period_end[0], period_end[1])

        # Excel
        st.subheader("3) Generar Diario (Apertura + 12 meses + Ajustes + Cierres + Mayor)")
        if st.button("Generar Diario"):
            xlsx_bytes = build_output_excel(
                empresa=empresa,
                opening_tuple=opening_tuple,
                month_tuples=month_tuples,
                adjust_blocks=adjust_blocks,
                cierre_resultado=cierre_resultado_lines,
                cierre_patrimonial=cierre_patrimonial_lines,
                period_end_date=period_end_date
            )
            st.success("‚úÖ Diario generado.")
            st.download_button(
                "‚¨áÔ∏è Descargar Diario (XLSX)",
                xlsx_bytes,
                "DiarioGenerado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error procesando el archivo: {e}")
else:
    st.info("Sub√≠ un Excel para comenzar.")
