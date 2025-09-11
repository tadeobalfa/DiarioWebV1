# -*- coding: utf-8 -*-
from __future__ import annotations
from typing import Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Paleta amarillo/naranja para detectar ajustes en As Aj
YELLOW_ORANGE_HEXES = {
    "FFFFFF00","FFFFF200","FFFFE699","FFFFF2CC","FFFFEB84","FFFFF4B6","FFFFF9CC",
    "FFFFC000","FFF4B183","FFF8CBAD","FFFFD965","FFFFB84D"
}
YELLOW_ORANGE_RGB = {h[-6:] for h in YELLOW_ORANGE_HEXES}

# Color separador celeste entre diarios y ajustes
SEPARATOR_COLOR = "FFB7DEE8"
BLUE_FILL = PatternFill(start_color=SEPARATOR_COLOR, end_color=SEPARATOR_COLOR, fill_type="solid")

# Bordes
THIN = Side(style="thin")
BORDER_BOTTOM = Border(bottom=THIN)  # bajo cabecera
BORDER_TOP = Border(top=THIN)        # sobre totales

def normalize_hex_from_cell(cell) -> Optional[str]:
    """
    Devuelve un ARGB/HEX en mayúsculas (p.ej. 'FFFFEB84') si puede inferirlo,
    tolerando que openpyxl devuelva objetos (RGB) en lugar de strings.
    """
    def extract_hex(color_obj):
        if color_obj is None:
            return None
        for attr in ("rgb", "value"):
            v = getattr(color_obj, attr, None)
            if v is None:
                continue
            # Desanidar si son objetos compuestos
            v = getattr(v, "rgb", v)
            v = getattr(v, "value", v)
            s = str(v).upper()  # robusto si no es string
            hexchars = "".join(ch for ch in s if ch in "0123456789ABCDEF")
            if len(hexchars) >= 6:
                return ("FF" + hexchars[-6:]) if len(hexchars) == 6 else hexchars
        return None

    for color in (
        getattr(cell.fill, "fgColor", None),
        getattr(cell.fill, "start_color", None),
        getattr(cell.fill, "end_color", None),
    ):
        hx = extract_hex(color)
        if hx:
            return hx
    return None

def get_top_left_cell(ws, cell):
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return cell

def is_numeric(val) -> bool:
    try:
        if val is None:
            return False
        float(str(val).strip().replace(",", "."))
        return True
    except:
        return False

def normalize_value(x):
    """Convierte a número (acepta 1.234,56). Si falla, 0.0"""
    if x is None or (isinstance(x, str) and x.strip() == ""):
        return 0.0
    try:
        s = str(x).replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        try:
            return float(x)
        except Exception:
            return 0.0

def coerce_num(x):
    """Solo para totales (si viene como string), tolerante a formatos ES."""
    if x in (None, ""):
        return 0.0
    try:
        return float(str(x).replace(".", "").replace(",", "."))
    except Exception:
        try:
            return float(x)
        except Exception:
            return 0.0

class DiarioWriter:
    # Columnas formato "Bien Verde"
    COL_A, COL_B, COL_C, COL_D, COL_E, COL_F = 1, 2, 3, 4, 5, 6

    def __init__(self, year: int):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = f"DIARIO {year}"

    def append_row(self, values: Dict[int, object]) -> int:
        r = self.ws.max_row + 1
        wrote = False
        for c, v in values.items():
            self.ws.cell(r, c, v)
            wrote = True
        if not wrote:
            self.ws.cell(r, self.COL_A, "")  # fila en blanco real
        return r

    def header(self, nro: int):
        row_idx = self.append_row({self.COL_A: "Fecha", self.COL_B: nro, self.COL_D: "D", self.COL_E: "H"})
        for c in (self.COL_A, self.COL_B, self.COL_C, self.COL_D, self.COL_E):
            self.ws.cell(row_idx, c).border = BORDER_BOTTOM

    def totales(self, titulo: str, total_d: float, total_h: float):
        row_idx = self.append_row({self.COL_A: titulo, self.COL_D: total_d, self.COL_E: total_h, self.COL_F: 0})
        for c in (self.COL_A, self.COL_B, self.COL_C, self.COL_D, self.COL_E, self.COL_F):
            self.ws.cell(row_idx, c).border = BORDER_TOP

    def blank_row(self):
        self.append_row({})

    def separator_blue(self):
        r = self.append_row({})
        for c in (self.COL_A, self.COL_B, self.COL_C, self.COL_D, self.COL_E, self.COL_F):
            self.ws.cell(r, c).fill = BLUE_FILL

    def save(self, path: str):
        widths = {1: 16, 2: 40, 3: 4, 4: 14, 5: 14, 6: 10}
        for c, w in widths.items():
            self.ws.column_dimensions[get_column_letter(c)].width = w
        self.wb.save(path)
