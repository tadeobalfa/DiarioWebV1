# -*- coding: utf-8 -*-
from __future__ import annotations
from typing import List, Tuple, Optional
from openpyxl import load_workbook
from .writer import (
    YELLOW_ORANGE_HEXES, YELLOW_ORANGE_RGB,
    normalize_hex_from_cell, get_top_left_cell, is_numeric
)

AJ_SHEET_CANDIDATES = ["As Aj", "AS AJ", "AS AJ.", "AS AJUSTE", "AS AJUSTES"]

def marker_is_highlighted(ws, row_idx: int) -> bool:
    """True si en la fila hay un NÚMERO de asiento pintado (A o B) en amarillo/naranja."""
    for col in (1, 2):  # A o B
        cell = ws.cell(row_idx, col)
        tl = get_top_left_cell(ws, cell)
        val = cell.value if cell.value is not None else tl.value
        hx1 = normalize_hex_from_cell(tl)
        hx2 = normalize_hex_from_cell(cell)
        hx = hx1 or hx2
        if is_numeric(val) and hx and ((hx in YELLOW_ORANGE_HEXES) or (hx[-6:] in YELLOW_ORANGE_RGB)):
            return True
    return False

def read_asaj_blocks(input_path: str, sheet_name: Optional[str]) -> List[Tuple[str, List[Tuple[str, object, object]]]]:
    """
    Lee la hoja 'As Aj' y devuelve:
      [(titulo, [(cuenta, debe_raw, haber_raw), ...]), ...]
    Copia Debe/Haber tal cual (raw) para que en el diario se vean igual.
    """
    wb = load_workbook(input_path, data_only=True)

    # Selección de hoja
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = None
        for c in AJ_SHEET_CANDIDATES:
            if c in wb.sheetnames:
                ws = wb[c]; break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

    blocks: List[Tuple[str, List[Tuple[str, object, object]]]] = []
    current: List[Tuple[str, object, object]] = []
    title: Optional[str] = None
    inside = False  # estamos dentro de un bloque entre dos marcadores pintados

    # Encabezado en fila 1: A=Cuenta, B=Nombre, C=Debe, D=Haber
    for r in range(2, ws.max_row + 1):
        # ¿Comienza un nuevo bloque?
        if marker_is_highlighted(ws, r):
            if inside and current:
                blocks.append((title or "As Ajuste", current))
            inside = True
            current = []
            title = None
            continue

        if inside:
            cuenta = ws.cell(r, 1).value
            nombre = ws.cell(r, 2).value
            debe = ws.cell(r, 3).value
            haber = ws.cell(r, 4).value

            if (cuenta is None and nombre is None and debe is None and haber is None):
                continue

            if title is None and nombre not in (None, ""):
                title = str(nombre).strip()

            if (cuenta is None or str(cuenta).strip() == "") and (debe in (None, "")) and (haber in (None, "")):
                continue

            cuenta_str = "" if cuenta is None else str(cuenta).strip()
            current.append((cuenta_str, debe, haber))

    if inside and current:
        blocks.append((title or "As Ajuste", current))

    return blocks
